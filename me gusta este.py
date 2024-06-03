from flask import Flask, request, render_template, send_file, redirect, url_for
from werkzeug.utils import secure_filename
import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import io

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def guardar_en_excel(datos, archivo_excel):
    if os.path.isfile(archivo_excel):
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_base, extension = os.path.splitext(archivo_excel)
        nuevo_archivo_excel = f"{nombre_base}_{fecha_hora}{extension}"
        archivo_excel = nuevo_archivo_excel

    df = pd.DataFrame(datos, columns=['Nombre', 'Perforacion', 'Joya', 'Suero/Cadena'])
    df['Comisión Perforacion'] = df['Perforacion'] * 0.30
    df['Comisión Joya'] = df['Joya'] * 0.25
    df['Comisión Suero/Cadena'] = df['Suero/Cadena'] * 0.15

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Ajustar la referencia de la tabla
    max_row = len(df) + 1
    max_col = 7
    table_ref = f"A1:G{max_row}"

    tab = Table(displayName="Table1", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Crear la hoja de resumen
    ws_summary = wb.create_sheet(title="Resumen")

    # Escribir los encabezados en la hoja de resumen
    headers = ['Perforacion', 'Joya', 'Suero/Cadena', 'Comisión Perforacion', 'Comisión Joya', 'Comisión Suero/Cadena']
    ws_summary.append([''] + headers)

    # Escribir las fórmulas para sumar cada columna
    for col_num, header in enumerate(headers, start=2):
        col_letter = chr(64 + col_num)
        formula = f"=SUBTOTAL(9,Datos!{col_letter}2:{col_letter}{max_row})"
        ws_summary.cell(row=2, column=col_num, value=formula)

    # Guardar el workbook en un objeto de bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def procesar_chat(texto_chat):
    texto_chat = texto_chat.replace(' ', '')  # Eliminar todos los espacios
    regex = r"([\w\s]+)-(\d+[\.,]?\d*)[-](\d+[\.,]?\d*)[-]?(\d+[\.,]?\d*)?"
    matches = re.findall(regex, texto_chat)
    return matches

@app.route('/pircing_max/calcula_comision', methods=['GET', 'POST'])
def calcula_comision():
    if request.method == 'POST':
        input_text = request.form.get('inputText')
        if input_text:
            datos_procesados = procesar_chat(input_text)
            datos_procesados = [(nombre.strip(), float(v1.replace('.', '').replace(',', '.')), float(v2.replace('.', '').replace(',', '.')), float(v3.replace('.', '').replace(',', '.') if v3 else 0)) for nombre, v1, v2, v3 in datos_procesados]
            output = guardar_en_excel(datos_procesados, 'comisiones.xlsx')
            return send_file(output, as_attachment=True, download_name='comisiones.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('calculadora.html')

if __name__ == '__main__':
    app.run(debug=True)
